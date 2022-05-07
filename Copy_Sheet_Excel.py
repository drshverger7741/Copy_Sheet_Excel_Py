import configparser
import os

import openpyxl
import xlsxwriter
import xlwings as xw

config = configparser.ConfigParser()
print(config.read('settings.ini'))

# удаление старого финального файла при наличии
final_file = config.get('final_name_file',
                        'final_file')
if os.path.isfile(final_file):
    os.remove(final_file)
    print("success")
else:
    print("File doesn't exists!")

# # получение списка файлов эксель
# cwd = os.getcwd()
# fileExt = r".xlsx"
# fileExt1 = r".xls"
# # List all files and directories in current directory
# files_exel = [os.path.join(cwd, _) for _ in os.listdir(cwd) if _.endswith(fileExt)]
# files_exels = [os.path.join(cwd, _) for _ in os.listdir(cwd) if _.endswith(fileExt1)]
# files_list = files_exel + files_exels
# print(files_list)

# создание финального файла
workbook = xlsxwriter.Workbook(final_file)
worksheet = workbook.add_worksheet()
workbook.close()

# копирование листов
xw.App(visible=config.getboolean('final_name_file',
                                 'visible'))
wb2 = xw.Book(final_file,
              update_links=config.getboolean('final_name_file',
                                             'update_links'),
              notify=config.getboolean('final_name_file',
                                       'notify'))
for key, value in config['files_sheet'].items():
    path1 = key
    wb1 = xw.Book(path1,
                  update_links=config.getboolean('every_file',
                                                 'update_links'),
                  notify=config.getboolean('every_file',
                                           'notify'))
    ws1 = wb1.sheets(value)
    ws1.api.Copy(Before=wb2.sheets(1).api)
    wb2.save()
    wb1.close()
wb2.app.quit()
# сортировка листов
# wb2 = xw.Book(final_file,
#               update_links=config.getboolean('final_name_file',
#                                              'update_links'),
#               notify=config.getboolean('final_name_file',
#                                        'notify'))
# list_poriadok = config['files_sheet'].items()
# print(list_poriadok)
# for worksheet in wb2.sheets:
#     if worksheet.Name == list_poriadok[-1]:
#         del list_poriadok[-1]
#         worksheet.Move(Before=wb2.sheets(1))
#         wb2.save()
# wb2.close()
# wb2.app.quit()

# удаление ненужных листов
wb = openpyxl.load_workbook(final_file)
sheets = wb.sheetnames
print(sheets)
for sheet in sheets:
    if sheet not in config['files_sheet'].values():
        pfd = wb[sheet]
        wb.remove(pfd)
wb.save(final_file)

# сортировка листов
# wb = xw.Book("test.xlsx")
# ws1 = wb.sheets['Sheet1']
# ws3 = wb.sheets['Sheet3']
# ws1.api.Move(None, After=ws3.api)

wb2 = xw.Book(final_file,
              update_links=config.getboolean('final_name_file',
                                             'update_links'),
              notify=config.getboolean('final_name_file',
                                       'notify'))
list_poriadok = config.get('list_poriadok', 'list_poriadok')
List = list_poriadok.split(',')
print(List)
for worksheet in wb2.sheets:
    if worksheet.Name == List[-1]:
        worksheet.Move(Before=wb2.sheets(1))
        del List[-1]
    wb2.save()
wb2.close()
wb2.app.quit()
