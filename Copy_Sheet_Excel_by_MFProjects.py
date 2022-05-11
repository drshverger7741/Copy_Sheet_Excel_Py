import configparser
import logging
import os
import shutil

import openpyxl
import xlwings as xw

config = configparser.ConfigParser()
config.read('settings.ini', encoding='utf-8')
logging.basicConfig(filename=config.get('debug',
                                        'log_file'),
                    filemode="w",
                    level=logging.INFO,
                    format="%(asctime)s - %(levelname)s - %(funcName)s: %(lineno)d - %(message)s")
logging.info('Получение данных настройки из settings.ini')
final_file = config.get('final_name_file',
                        'final_file')

# Если хотим копировать первые листы всех файлов в директории
if config.getboolean('final_name_file', 'copy_first_sheet_allfiles'):
    if os.path.isfile(final_file):
        os.remove(final_file)
    logging.info('Будет произведено копирование всех листов из директории в один новый файл')
fileDir = os.getcwd()
fileExtLISTx = [os.path.join(fileDir, _) for _ in os.listdir(fileDir) if _.endswith(r".xlsx")]
fileExtLIS = [os.path.join(fileDir, _) for _ in os.listdir(fileDir) if _.endswith(r".xls")]
fileExtLIST = fileExtLISTx + fileExtLIS
print(fileExtLIST)

# создание финального файла
wb = openpyxl.Workbook()
wb.save('1.xlsx')
if config.getboolean('final_name_file', 'create_new_file'):
    shutil.copyfile('1.xlsx', final_file)
    logging.info("Создан новый файл: " + final_file)
elif os.path.isfile(final_file):
    logging.info(
        "Новый файл не будет создан все копирование будет в существующий: " + final_file +
        ". Пожалуйста не забудьте в настройках settings.ini указать какие лиcты необходимо сохранить"
        " в финальном файле после копирования")
else:
    logging.info(
        "Файла с таким именем не найдено(поэтому будет создан новый пустой и туда будут скопированы листы: "
        + final_file)
    shutil.copyfile('1.xlsx', final_file)
os.remove('1.xlsx')

# копирование листов
logging.info("Началось копирование листов в " + final_file)
app = xw.App(visible=config.getboolean('final_name_file',
                                       'visible'))
wb2 = xw.Book(final_file,
              update_links=config.getboolean('final_name_file',
                                             'update_links'),
              notify=config.getboolean('final_name_file',
                                       'notify'))
if config.getboolean('final_name_file', 'copy_first_sheet_allfiles'):
    for all_xl in fileExtLIST:
        wb1 = xw.Book(all_xl,
                      update_links=config.getboolean('every_file',
                                                     'update_links'),
                      notify=config.getboolean('every_file',
                                               'notify'))
        ws1 = wb1.sheets(config.getint('final_name_file', 'copy_first_sheet_allfiles_number'))
        ws1.api.Copy(Before=None, After=wb2.sheets(wb2.sheets.count).api)
        wb2.save()
        logging.info("Лист из книги " + all_xl + " успешно скопирован в " + final_file)
        wb1.close()
    wb2.app.quit()
else:
    for key, value in config['files_sheet'].items():
        path1 = key
        wb1 = xw.Book(path1,
                      update_links=config.getboolean('every_file',
                                                     'update_links'),
                      notify=config.getboolean('every_file',
                                               'notify'))
        ws1 = wb1.sheets(value)
        ws1.api.Copy(Before=None, After=wb2.sheets(wb2.sheets.count).api)
        wb2.save()
        logging.info("Лист " + value + " из книги " + key + " успешно скопирован в " + final_file)
        wb1.close()
    wb2.app.quit()

# удаление ненужных листов
wb = openpyxl.load_workbook(final_file)
sheets = wb.sheetnames
list_sheet_final_file = config.get('final_name_file', 'list_sheet_final_file').split(",")
for sheet in sheets:
    if config.getboolean('final_name_file', 'create_new_file') and sheet == "Sheet":
        pfd = wb[sheet]
        wb.remove(pfd)
    if not config.getboolean('final_name_file',
                             'create_new_file') and sheet not in list_sheet_final_file:
        pfd = wb[sheet]
        wb.remove(pfd)
        logging.info("Лист: " + sheet + " не из списка list_sheet_final_file удален успешно")
sheets = wb.sheetnames
print(sheets)
wb.save(final_file)
logging.info("Копирование листов успешно выполнено!")
